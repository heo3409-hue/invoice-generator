#!/usr/bin/env python3
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter
import io, os
from datetime import datetime

app = Flask(__name__)
CORS(app)
BASE = os.path.dirname(os.path.abspath(__file__))

TEMPLATES = {
    'china':  os.path.join(BASE, 'template_china.xlsx'),   # TS-260119 (2품목: G23/H23, G25/H25)
    'india':  os.path.join(BASE, 'template_india.xlsx'),   # TS-260501 (2품목: G24/H24, G25/H25)
    'nasn':   os.path.join(BASE, 'template_nasn.xlsx'),    # TS-260511 (1품목: G24/H24)
    'mexico': os.path.join(BASE, 'template_mexico.xlsx'),  # TS-260504 (2품목: G24/H24, G26/H26)
}

def fmt_ordinal(date_str):
    d = datetime.strptime(date_str, '%Y-%m-%d')
    day = d.day
    suf = {1:'st',2:'nd',3:'rd'}.get(day if day<20 else day%10, 'th')
    return f"{day}{suf}.{d.strftime('%b')}.{d.year}"

def fmt_upper(date_str):
    d = datetime.strptime(date_str, '%Y-%m-%d')
    return f"{d.day}-{d.strftime('%b').upper()}-{d.year}"

def safe_write(ws, ref, value):
    col_str, row = coordinate_from_string(ref)
    col = column_index_from_string(col_str)
    for m in ws.merged_cells.ranges:
        if m.min_row <= row <= m.max_row and m.min_col <= col <= m.max_col:
            master = f"{get_column_letter(m.min_col)}{m.min_row}"
            if ref != master:
                return
    try:
        ws[ref].value = value
    except AttributeError:
        pass

@app.route('/generate', methods=['POST'])
def generate():
    data = request.json
    cid = data.get('client_id')
    if cid not in TEMPLATES:
        return jsonify({'error': f'Unknown client: {cid}'}), 400

    wb = load_workbook(TEMPLATES[cid])
    items = data.get('items', [])
    inv_no = data.get('inv_no', 'TS-000000-01')
    date_str = data.get('date', '2026-01-01')

    if cid == 'china':
        ws = wb['Invoice']
        safe_write(ws, 'F3', f"INVOICE NO : {inv_no}")
        safe_write(ws, 'F4', f"INVOICE DATE : {fmt_ordinal(date_str)}")
        if len(items) >= 1:
            safe_write(ws, 'G23', items[0]['qty'])
            safe_write(ws, 'H23', items[0]['uprice'])
        if len(items) >= 2:
            safe_write(ws, 'G25', items[1]['qty'])
            safe_write(ws, 'H25', items[1]['uprice'])

    elif cid == 'india':
        ws = wb['INVOICE']
        safe_write(ws, 'F3', f"INVOICE NO : {inv_no}")
        safe_write(ws, 'F4', f"INVOICE DATE : {fmt_ordinal(date_str)}")
        if len(items) >= 1:
            safe_write(ws, 'G24', items[0]['qty'])
            safe_write(ws, 'H24', items[0]['uprice'])
        if len(items) >= 2:
            safe_write(ws, 'G25', items[1]['qty'])
            safe_write(ws, 'H25', items[1]['uprice'])

    elif cid == 'nasn':
        ws = wb['INVOICE ']
        safe_write(ws, 'F3', f"INVOICE NO : {inv_no}")
        safe_write(ws, 'F4', f"INVOICE DATE : {fmt_upper(date_str)}")
        if len(items) >= 1:
            safe_write(ws, 'G24', items[0]['qty'])
            safe_write(ws, 'H24', items[0]['uprice'])

    elif cid == 'mexico':
        ws = wb['INVOICE']
        safe_write(ws, 'F3', f"INVOICE NO : {inv_no}")
        safe_write(ws, 'F4', f"INVOICE DATE : {fmt_ordinal(date_str)}")
        if len(items) >= 1:
            safe_write(ws, 'G24', items[0]['qty'])
            safe_write(ws, 'H24', items[0]['uprice'])
        if len(items) >= 2:
            safe_write(ws, 'G26', items[1]['qty'])
            safe_write(ws, 'H26', items[1]['uprice'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"{inv_no}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/health')
def health():
    return jsonify({'status': 'ok'})

if __name__ == '__main__':
    print("✓ Invoice Generator Server 시작")
    print("✓ http://localhost:5050")
    app.run(host='0.0.0.0', port=5050, debug=False)
